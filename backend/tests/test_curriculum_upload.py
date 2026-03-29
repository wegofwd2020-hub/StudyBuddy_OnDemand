"""
tests/test_curriculum_upload.py

Tests for Phase 8 curriculum upload + pipeline endpoints:
  GET  /api/v1/curriculum/template
  POST /api/v1/curriculum/upload
  POST /api/v1/curriculum/upload/xlsx
  POST /api/v1/curriculum/pipeline/trigger
  GET  /api/v1/curriculum/pipeline/{job_id}/status

Also unit-tests for:
  - parse_xlsx (structural error handling)
  - promote_student_grades Celery task (no-op on wrong date, runs on match)
"""

from __future__ import annotations

import io
import json
import uuid
from unittest.mock import AsyncMock, MagicMock, patch

import pytest
from httpx import AsyncClient

from tests.helpers.token_factory import make_teacher_token

# ── Helpers ───────────────────────────────────────────────────────────────────

_VALID_UNITS = [
    {
        "subject": "Mathematics",
        "unit_name": "Linear Equations",
        "unit_id": "MATH-001",
        "objectives": ["Solve equations", "Graph functions"],
        "has_lab": False,
        "lab_description": None,
    },
    {
        "subject": "Science",
        "unit_name": "Measuring Density",
        "unit_id": "SCI-001",
        "objectives": ["Apply density formula", "Use lab equipment safely"],
        "has_lab": True,
        "lab_description": "Measure density of solids using a balance.",
    },
]


async def _register_school(client: AsyncClient, suffix: str = "") -> dict:
    """Register a school and return {school_id, teacher_id, access_token}."""
    r = await client.post("/api/v1/schools/register", json={
        "school_name": f"Upload Test School{suffix}",
        "contact_email": f"upload{suffix}@testschool.example.com",
        "country": "US",
    })
    assert r.status_code == 201, r.text
    return r.json()


def _make_xlsx_bytes(grade: int = 8) -> bytes:
    """Generate a minimal valid XLSX for the given grade."""
    from src.curriculum.upload_service import build_xlsx_template
    return build_xlsx_template(grade)


# ── GET /curriculum/template ──────────────────────────────────────────────────

@pytest.mark.asyncio
async def test_download_template_returns_xlsx(client: AsyncClient):
    """Template endpoint returns xlsx bytes."""
    r = await client.get("/api/v1/curriculum/template?grade=8")
    assert r.status_code == 200
    assert "spreadsheetml" in r.headers["content-type"]
    assert len(r.content) > 100


@pytest.mark.asyncio
async def test_download_template_invalid_grade_returns_400(client: AsyncClient):
    """Grade out of range returns 400."""
    r = await client.get("/api/v1/curriculum/template?grade=3")
    assert r.status_code == 400


# ── POST /curriculum/upload ───────────────────────────────────────────────────

@pytest.mark.asyncio
async def test_upload_curriculum_json_succeeds(client: AsyncClient):
    """Valid JSON curriculum upload returns 201 with curriculum_id."""
    school = await _register_school(client, "-upload-json")
    headers = {"Authorization": f"Bearer {school['access_token']}"}

    payload = {
        "grade": 8,
        "year": 2026,
        "name": "Test Grade 8 STEM",
        "units": _VALID_UNITS,
    }
    r = await client.post("/api/v1/curriculum/upload", json=payload, headers=headers)
    assert r.status_code == 201, r.text
    data = r.json()
    assert "curriculum_id" in data
    assert data["unit_count"] == 2
    assert data["errors"] == []


@pytest.mark.asyncio
async def test_upload_curriculum_json_requires_auth(client: AsyncClient):
    """Upload without JWT returns 401."""
    payload = {"grade": 8, "year": 2026, "name": "Test", "units": _VALID_UNITS}
    r = await client.post("/api/v1/curriculum/upload", json=payload)
    assert r.status_code == 401


@pytest.mark.asyncio
async def test_upload_curriculum_json_validation_error_returns_400(client: AsyncClient):
    """Units missing required objectives return 400 with structured errors."""
    school = await _register_school(client, "-val-error")
    headers = {"Authorization": f"Bearer {school['access_token']}"}

    bad_units = [
        {
            "subject": "Math",
            "unit_name": "Bad Unit",
            "unit_id": "BAD-001",
            "objectives": ["Only one"],  # less than 2 required
            "has_lab": False,
            "lab_description": None,
        }
    ]
    payload = {"grade": 8, "year": 2026, "name": "Bad Upload", "units": bad_units}
    r = await client.post("/api/v1/curriculum/upload", json=payload, headers=headers)
    assert r.status_code == 400
    data = r.json()
    assert data["error"] == "validation_error"
    assert len(data["errors"]) >= 1


@pytest.mark.asyncio
async def test_upload_curriculum_json_has_lab_no_description_returns_400(client: AsyncClient):
    """has_lab=True without lab_description triggers validation error."""
    school = await _register_school(client, "-lab-val")
    headers = {"Authorization": f"Bearer {school['access_token']}"}

    bad_units = [
        {
            "subject": "Science",
            "unit_name": "Lab Unit",
            "unit_id": "LAB-001",
            "objectives": ["Obj one", "Obj two"],
            "has_lab": True,
            "lab_description": None,
        }
    ]
    payload = {"grade": 8, "year": 2026, "name": "Lab Upload", "units": bad_units}
    r = await client.post("/api/v1/curriculum/upload", json=payload, headers=headers)
    assert r.status_code == 400
    errors = r.json()["errors"]
    assert any(e["field"] == "Lab Description" for e in errors)


@pytest.mark.asyncio
async def test_upload_curriculum_json_duplicate_unit_code_returns_400(client: AsyncClient):
    """Duplicate unit codes within one upload return 400."""
    school = await _register_school(client, "-dup-code")
    headers = {"Authorization": f"Bearer {school['access_token']}"}

    dup_units = [
        {
            "subject": "Math",
            "unit_name": "Unit A",
            "unit_id": "DUP-001",
            "objectives": ["Obj one", "Obj two"],
            "has_lab": False,
            "lab_description": None,
        },
        {
            "subject": "Science",
            "unit_name": "Unit B",
            "unit_id": "DUP-001",  # duplicate
            "objectives": ["Obj one", "Obj two"],
            "has_lab": False,
            "lab_description": None,
        },
    ]
    payload = {"grade": 8, "year": 2026, "name": "Dup Upload", "units": dup_units}
    r = await client.post("/api/v1/curriculum/upload", json=payload, headers=headers)
    assert r.status_code == 400
    errors = r.json()["errors"]
    assert any("Duplicate" in e["message"] for e in errors)


# ── POST /curriculum/upload/xlsx ──────────────────────────────────────────────

@pytest.mark.asyncio
async def test_upload_xlsx_template_succeeds(client: AsyncClient):
    """XLSX template (with sample rows) can be round-tripped through the upload."""
    school = await _register_school(client, "-xlsx")
    headers = {"Authorization": f"Bearer {school['access_token']}"}

    xlsx_bytes = _make_xlsx_bytes(8)
    r = await client.post(
        "/api/v1/curriculum/upload/xlsx",
        headers=headers,
        files={"file": ("curriculum.xlsx", xlsx_bytes,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        params={"grade": 8, "year": 2026, "name": "XLSX Test"},
    )
    assert r.status_code == 201, r.text
    data = r.json()
    assert "curriculum_id" in data
    assert data["unit_count"] >= 2


@pytest.mark.asyncio
async def test_upload_xlsx_requires_auth(client: AsyncClient):
    """XLSX upload without JWT returns 401."""
    xlsx_bytes = _make_xlsx_bytes(8)
    r = await client.post(
        "/api/v1/curriculum/upload/xlsx",
        files={"file": ("curriculum.xlsx", xlsx_bytes,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        params={"grade": 8},
    )
    assert r.status_code == 401


@pytest.mark.asyncio
async def test_upload_xlsx_wrong_grade_returns_400(client: AsyncClient):
    """Grade out of range returns 400 without parsing the file."""
    school = await _register_school(client, "-xlsx-grade")
    headers = {"Authorization": f"Bearer {school['access_token']}"}

    xlsx_bytes = _make_xlsx_bytes(8)
    r = await client.post(
        "/api/v1/curriculum/upload/xlsx",
        headers=headers,
        files={"file": ("curriculum.xlsx", xlsx_bytes,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        params={"grade": 3},
    )
    assert r.status_code == 400


@pytest.mark.asyncio
async def test_upload_xlsx_wrong_sheet_name_returns_400(client: AsyncClient):
    """XLSX missing the expected sheet returns 400 with structured error."""
    import openpyxl
    school = await _register_school(client, "-xlsx-sheet")
    headers = {"Authorization": f"Bearer {school['access_token']}"}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "WrongSheet"
    ws.append(["Subject", "Unit Name", "Objectives"])
    ws.append(["Math", "Some Unit", "Obj one|Obj two"])
    buf = io.BytesIO()
    wb.save(buf)
    xlsx_bytes = buf.getvalue()

    r = await client.post(
        "/api/v1/curriculum/upload/xlsx",
        headers=headers,
        files={"file": ("curriculum.xlsx", xlsx_bytes,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        params={"grade": 8},
    )
    assert r.status_code == 400
    data = r.json()
    assert data["error"] == "validation_error"


# ── parse_xlsx unit tests ─────────────────────────────────────────────────────

def test_parse_xlsx_missing_required_column():
    """parse_xlsx returns error when required column is missing."""
    import openpyxl
    from src.curriculum.upload_service import parse_xlsx

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Grade_8"
    ws.append(["Subject", "Unit Name"])  # missing Objectives
    ws.append(["Math", "Some Unit"])
    buf = io.BytesIO()
    wb.save(buf)

    units, errors = parse_xlsx(buf.getvalue(), 8)
    assert units == []
    assert any("Objectives" in e["message"] for e in errors)


def test_parse_xlsx_wrong_sheet_name():
    """parse_xlsx returns error when sheet name doesn't match grade."""
    import openpyxl
    from src.curriculum.upload_service import parse_xlsx

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Grade_9"
    ws.append(["Subject", "Unit Name", "Objectives"])
    buf = io.BytesIO()
    wb.save(buf)

    units, errors = parse_xlsx(buf.getvalue(), 8)
    assert units == []
    assert any("Grade_8" in e["message"] for e in errors)


def test_parse_xlsx_valid_data():
    """parse_xlsx correctly parses a well-formed sheet."""
    import openpyxl
    from src.curriculum.upload_service import parse_xlsx

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Grade_8"
    ws.append(["Subject", "Unit Name", "Unit Code", "Objectives", "Has Lab", "Lab Description"])
    ws.append(["Mathematics", "Algebra", "MATH-001", "Obj one|Obj two", "No", ""])
    ws.append(["Science", "Density Lab", "SCI-001",
               "Apply formula|Use equipment safely", "Yes", "Measure density."])
    buf = io.BytesIO()
    wb.save(buf)

    units, errors = parse_xlsx(buf.getvalue(), 8)
    assert errors == []
    assert len(units) == 2
    assert units[0]["subject"] == "Mathematics"
    assert units[0]["objectives"] == ["Obj one", "Obj two"]
    assert units[1]["has_lab"] is True
    assert units[1]["lab_description"] == "Measure density."


# ── POST /curriculum/pipeline/trigger ─────────────────────────────────────────

@pytest.mark.asyncio
async def test_pipeline_trigger_returns_202_with_job_id(client: AsyncClient):
    """Triggering a pipeline returns 202 with a job_id."""
    school = await _register_school(client, "-pipeline")
    headers = {"Authorization": f"Bearer {school['access_token']}"}

    upload_r = await client.post(
        "/api/v1/curriculum/upload",
        json={"grade": 8, "year": 2026, "name": "Pipeline Test", "units": _VALID_UNITS},
        headers=headers,
    )
    assert upload_r.status_code == 201, upload_r.text
    curriculum_id = upload_r.json()["curriculum_id"]

    with patch("src.auth.tasks.celery_app") as mock_celery:
        mock_celery.send_task = MagicMock()
        r = await client.post(
            "/api/v1/curriculum/pipeline/trigger",
            json={"curriculum_id": curriculum_id, "langs": "en", "force": False},
            headers=headers,
        )
    assert r.status_code == 202, r.text
    data = r.json()
    assert "job_id" in data
    assert data["status"] == "queued"


@pytest.mark.asyncio
async def test_pipeline_trigger_nonexistent_curriculum_returns_404(client: AsyncClient):
    """Triggering pipeline for a non-existent curriculum_id returns 404."""
    school = await _register_school(client, "-pipeline-404")
    headers = {"Authorization": f"Bearer {school['access_token']}"}

    with patch("src.auth.tasks.celery_app") as mock_celery:
        mock_celery.send_task = MagicMock()
        r = await client.post(
            "/api/v1/curriculum/pipeline/trigger",
            json={"curriculum_id": str(uuid.uuid4()), "langs": "en", "force": False},
            headers=headers,
        )
    assert r.status_code == 404


@pytest.mark.asyncio
async def test_pipeline_trigger_requires_auth(client: AsyncClient):
    """Pipeline trigger without JWT returns 401."""
    r = await client.post(
        "/api/v1/curriculum/pipeline/trigger",
        json={"curriculum_id": str(uuid.uuid4()), "langs": "en", "force": False},
    )
    assert r.status_code == 401


# ── GET /curriculum/pipeline/{job_id}/status ─────────────────────────────────

@pytest.mark.asyncio
async def test_pipeline_status_returns_job_state(client: AsyncClient):
    """Polling an existing job_id returns its state."""
    school = await _register_school(client, "-status")
    headers = {"Authorization": f"Bearer {school['access_token']}"}

    upload_r = await client.post(
        "/api/v1/curriculum/upload",
        json={"grade": 8, "year": 2026, "name": "Status Test", "units": _VALID_UNITS},
        headers=headers,
    )
    assert upload_r.status_code == 201
    curriculum_id = upload_r.json()["curriculum_id"]

    with patch("src.auth.tasks.celery_app") as mock_celery:
        mock_celery.send_task = MagicMock()
        trigger_r = await client.post(
            "/api/v1/curriculum/pipeline/trigger",
            json={"curriculum_id": curriculum_id, "langs": "en", "force": False},
            headers=headers,
        )
    assert trigger_r.status_code == 202
    job_id = trigger_r.json()["job_id"]

    r = await client.get(
        f"/api/v1/curriculum/pipeline/{job_id}/status",
        headers=headers,
    )
    assert r.status_code == 200, r.text
    data = r.json()
    assert data["job_id"] == job_id
    assert data["status"] == "queued"
    assert "total" in data
    assert "built" in data
    assert "progress_pct" in data


@pytest.mark.asyncio
async def test_pipeline_status_nonexistent_job_returns_404(client: AsyncClient):
    """Polling an unknown job_id returns 404."""
    school = await _register_school(client, "-status-404")
    headers = {"Authorization": f"Bearer {school['access_token']}"}

    r = await client.get(
        f"/api/v1/curriculum/pipeline/{uuid.uuid4()}/status",
        headers=headers,
    )
    assert r.status_code == 404


@pytest.mark.asyncio
async def test_pipeline_status_requires_auth(client: AsyncClient):
    """Status endpoint without JWT returns 401."""
    r = await client.get(f"/api/v1/curriculum/pipeline/{uuid.uuid4()}/status")
    assert r.status_code == 401


# ── promote_student_grades task (unit tests) ──────────────────────────────────

def test_promote_student_grades_no_op_on_wrong_date():
    """Task returns early if today does not match GRADE_PROMOTION_DATE."""
    from src.auth.tasks import promote_student_grades

    with patch("src.auth.tasks.settings") as mock_settings:
        mock_settings.GRADE_PROMOTION_DATE = "01-01"  # not today (2026-03-26)
        mock_settings.DATABASE_URL = "postgresql://x/x"
        with patch("src.auth.tasks._run_async") as mock_run_async:
            promote_student_grades()
            mock_run_async.assert_not_called()


def test_promote_student_grades_skips_when_no_date_configured():
    """Task is a no-op when GRADE_PROMOTION_DATE is not set."""
    from src.auth.tasks import promote_student_grades

    with patch("src.auth.tasks.settings") as mock_settings:
        mock_settings.GRADE_PROMOTION_DATE = None
        mock_settings.DATABASE_URL = "postgresql://x/x"
        with patch("src.auth.tasks._run_async") as mock_run_async:
            promote_student_grades()
            mock_run_async.assert_not_called()
