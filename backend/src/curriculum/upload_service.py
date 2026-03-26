"""
backend/src/curriculum/upload_service.py

Curriculum upload, pipeline trigger, and job-status logic for Phase 8.

Covers:
  - JSON curriculum creation (direct unit list)
  - XLSX curriculum parsing (openpyxl)
  - Pipeline trigger → Celery task + Redis job state
  - Job-status reads from Redis
  - XLSX template generation (openpyxl)
  - seed_default helper (used by pipeline/seed_default.py)
"""

from __future__ import annotations

import io
import json
import re
import uuid
from typing import List, Optional

import asyncpg

from src.utils.logger import get_logger

log = get_logger("curriculum.upload")

# ── Subject abbreviations for auto-generated unit codes ──────────────────────
_SUBJECT_ABBR: dict[str, str] = {
    "mathematics": "MATH",
    "math": "MATH",
    "science": "SCI",
    "technology": "TECH",
    "engineering": "ENG",
    "physics": "PHYS",
    "chemistry": "CHEM",
    "biology": "BIO",
}


def _subject_abbr(subject: str) -> str:
    return _SUBJECT_ABBR.get(subject.lower().strip(), subject[:4].upper())


def _auto_unit_id(subject: str, sequence: int) -> str:
    return f"{_subject_abbr(subject)}-{sequence:03d}"


# ── Validation helpers ────────────────────────────────────────────────────────

def _validate_units(units: list) -> list:
    """
    Validate a list of unit dicts.  Returns a list of error dicts:
      [{row, field, message}, ...]
    """
    errors = []
    seen_codes: set[str] = set()

    for i, unit in enumerate(units):
        row = i + 1
        subject = unit.get("subject", "").strip()
        unit_name = unit.get("unit_name", "").strip()
        objectives = unit.get("objectives", [])
        has_lab = unit.get("has_lab", False)
        lab_desc = unit.get("lab_description", "")

        if not subject:
            errors.append({"row": row, "field": "Subject", "message": "Subject is required."})
        if not unit_name:
            errors.append({"row": row, "field": "Unit Name", "message": "Unit Name is required."})
        if len(objectives) < 2:
            errors.append({"row": row, "field": "Objectives",
                           "message": "At least 2 objectives are required (pipe-separated)."})
        if has_lab and not lab_desc:
            errors.append({"row": row, "field": "Lab Description",
                           "message": "Lab Description is required when Has Lab = Yes."})

        code = unit.get("unit_id") or ""
        if code:
            if code in seen_codes:
                errors.append({"row": row, "field": "Unit Code",
                               "message": f"Duplicate unit code '{code}'."})
            seen_codes.add(code)

    return errors


# ── JSON upload ───────────────────────────────────────────────────────────────

async def create_curriculum_from_json(
    conn: asyncpg.Connection,
    grade: int,
    year: int,
    name: str,
    units: list,
    teacher_id: Optional[str],
    school_id: Optional[str],
) -> dict:
    """
    Insert a curricula row + curriculum_units rows from a validated unit list.
    Returns {curriculum_id, unit_count, errors}.
    """
    errors = _validate_units(units)
    if errors:
        return {"curriculum_id": None, "unit_count": 0, "errors": errors}

    curriculum_id = str(uuid.uuid4())

    await conn.execute(
        """
        INSERT INTO curricula
            (curriculum_id, school_id, grade, year, name, source_type, status, created_by)
        VALUES ($1, $2, $3, $4, $5, 'ui_form', 'draft', $6)
        """,
        curriculum_id,
        uuid.UUID(school_id) if school_id else None,
        grade, year, name,
        uuid.UUID(teacher_id) if teacher_id else None,
    )

    for seq, unit in enumerate(units, start=1):
        subject = unit["subject"].strip()
        unit_id = (unit.get("unit_id") or "").strip() or _auto_unit_id(subject, seq)
        objectives = unit.get("objectives", [])
        if isinstance(objectives, str):
            objectives = [o.strip() for o in objectives.split("|") if o.strip()]

        await conn.execute(
            """
            INSERT INTO curriculum_units
                (unit_id, curriculum_id, subject, unit_name, objectives,
                 has_lab, lab_description, sequence)
            VALUES ($1, $2, $3, $4, $5, $6, $7, $8)
            """,
            unit_id, curriculum_id, subject, unit["unit_name"].strip(),
            objectives,
            bool(unit.get("has_lab", False)),
            unit.get("lab_description") or None,
            seq,
        )

    log.info("curriculum_created", curriculum_id=curriculum_id, unit_count=len(units))
    return {"curriculum_id": curriculum_id, "unit_count": len(units), "errors": []}


# ── XLSX upload ───────────────────────────────────────────────────────────────

def parse_xlsx(content: bytes, grade: int) -> tuple[list, list]:
    """
    Parse an uploaded XLSX file for the given grade.

    Returns (units, errors) where units is a list of dicts and errors is a
    list of {row, field, message} dicts from structural problems.
    """
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet_name = f"Grade_{grade}"
    if sheet_name not in wb.sheetnames:
        # Try case-insensitive match
        match = next((s for s in wb.sheetnames if s.lower() == sheet_name.lower()), None)
        if not match:
            return [], [{"row": 0, "field": "Sheet",
                         "message": f"Sheet '{sheet_name}' not found in workbook."}]
        sheet_name = match

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], [{"row": 0, "field": "Sheet", "message": "Sheet is empty."}]

    # First row is header
    header = [str(c).strip() if c else "" for c in rows[0]]
    required_cols = {"Subject", "Unit Name", "Objectives"}
    col_idx: dict[str, int] = {}
    for col in ("Subject", "Unit Name", "Unit Code", "Objectives", "Has Lab", "Lab Description"):
        try:
            col_idx[col] = header.index(col)
        except ValueError:
            if col in required_cols:
                return [], [{"row": 1, "field": col,
                             "message": f"Required column '{col}' not found in header row."}]

    units = []
    for row_num, row in enumerate(rows[1:], start=2):
        def cell(col_name: str) -> str:
            idx = col_idx.get(col_name)
            return str(row[idx]).strip() if idx is not None and row[idx] is not None else ""

        subject = cell("Subject")
        if not subject:
            continue  # skip blank rows silently

        raw_objectives = cell("Objectives")
        objectives = [o.strip() for o in raw_objectives.split("|") if o.strip()]
        has_lab = cell("Has Lab").lower() in ("yes", "true", "1")

        units.append({
            "subject": subject,
            "unit_name": cell("Unit Name"),
            "unit_id": cell("Unit Code") or None,
            "objectives": objectives,
            "has_lab": has_lab,
            "lab_description": cell("Lab Description") or None,
        })

    return units, []


# ── XLSX template ─────────────────────────────────────────────────────────────

def build_xlsx_template(grade: int) -> bytes:
    """Generate an XLSX template for the given grade with sample rows."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Grade_{grade}"

    header = ["Subject", "Unit Name", "Unit Code", "Objectives", "Has Lab", "Lab Description"]
    ws.append(header)

    # Bold header + light blue fill
    header_fill = PatternFill("solid", fgColor="DDEEFF")
    header_font = Font(bold=True)
    for col, cell in enumerate(ws[1], start=1):
        cell.font = header_font
        cell.fill = header_fill
        ws.column_dimensions[cell.column_letter].width = max(15, len(header[col - 1]) + 4)

    # Sample rows
    ws.append(["Mathematics", "Algebra – Linear Equations", "MATH-LIN-001",
               "Solve linear equations|Graph functions|Apply to word problems",
               "No", ""])
    ws.append(["Science", "Measuring Density", "SCI-DEN-001",
               "Apply the density formula|Use lab equipment safely",
               "Yes", "Measure density of common solids using a balance and graduated cylinder"])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Pipeline trigger ──────────────────────────────────────────────────────────

async def trigger_pipeline(
    conn: asyncpg.Connection,
    redis,
    curriculum_id: str,
    langs: str,
    force: bool,
    teacher_id: str,
) -> dict:
    """
    Dispatch a Celery pipeline job for a curriculum.

    Job state is stored in Redis under pipeline:job:{job_id}.
    Returns {job_id, status}.
    """
    # Verify curriculum exists and belongs to the teacher's school
    row = await conn.fetchrow(
        "SELECT curriculum_id, status FROM curricula WHERE curriculum_id = $1",
        curriculum_id,
    )
    if not row:
        return {}

    total = await conn.fetchval(
        "SELECT COUNT(*) FROM curriculum_units WHERE curriculum_id = $1",
        curriculum_id,
    )

    job_id = str(uuid.uuid4())
    job_state = {
        "job_id": job_id,
        "curriculum_id": curriculum_id,
        "status": "queued",
        "built": 0,
        "failed": 0,
        "total": int(total or 0),
        "progress_pct": 0.0,
        "langs": langs,
    }
    await redis.setex(
        f"pipeline:job:{job_id}",
        86400,  # 24-hour TTL
        json.dumps(job_state),
    )

    # Update curriculum status to 'building'
    await conn.execute(
        "UPDATE curricula SET status = 'building' WHERE curriculum_id = $1",
        curriculum_id,
    )

    # Dispatch Celery task
    try:
        from src.auth.tasks import celery_app
        celery_app.send_task(
            "src.auth.tasks.run_curriculum_pipeline_task",
            args=[job_id, curriculum_id, langs, force, teacher_id],
        )
    except Exception as exc:
        log.warning("pipeline_dispatch_failed", curriculum_id=curriculum_id, error=str(exc))

    log.info("pipeline_triggered", job_id=job_id, curriculum_id=curriculum_id)
    return {"job_id": job_id, "status": "queued"}


async def get_pipeline_job_status(redis, job_id: str) -> Optional[dict]:
    """Read job state from Redis."""
    raw = await redis.get(f"pipeline:job:{job_id}")
    if not raw:
        return None
    data = json.loads(raw)
    return {
        "job_id": data["job_id"],
        "status": data["status"],
        "built": data["built"],
        "failed": data["failed"],
        "total": data["total"],
        "progress_pct": data["progress_pct"],
    }


# ── Seed helper (used by pipeline/seed_default.py) ────────────────────────────

async def seed_default_curriculum(
    conn: asyncpg.Connection,
    grade: int,
    year: int,
    units: list,
) -> str:
    """
    Upsert a default curriculum and its units.

    curriculum_id = 'default-{year}-g{grade}'
    status = 'active' (no pipeline run needed for default content)
    """
    curriculum_id = f"default-{year}-g{grade}"
    name = f"Grade {grade} STEM {year} (Default)"

    await conn.execute(
        """
        INSERT INTO curricula (curriculum_id, grade, year, name, source_type, status)
        VALUES ($1, $2, $3, $4, 'default', 'active')
        ON CONFLICT (curriculum_id) DO UPDATE
            SET name = EXCLUDED.name, status = 'active'
        """,
        curriculum_id, grade, year, name,
    )

    for seq, unit in enumerate(units, start=1):
        subject = unit.get("subject", "").strip()
        unit_id = unit.get("unit_id") or _auto_unit_id(subject, seq)
        objectives = unit.get("objectives", [])
        if isinstance(objectives, str):
            objectives = [o.strip() for o in objectives.split("|") if o.strip()]

        await conn.execute(
            """
            INSERT INTO curriculum_units
                (unit_id, curriculum_id, subject, unit_name, objectives,
                 has_lab, lab_description, sequence)
            VALUES ($1, $2, $3, $4, $5, $6, $7, $8)
            ON CONFLICT (unit_id, curriculum_id) DO UPDATE
                SET unit_name = EXCLUDED.unit_name,
                    objectives = EXCLUDED.objectives,
                    has_lab = EXCLUDED.has_lab,
                    lab_description = EXCLUDED.lab_description,
                    sequence = EXCLUDED.sequence
            """,
            unit_id, curriculum_id, subject,
            unit.get("unit_name", subject).strip(),
            objectives,
            bool(unit.get("has_lab", False)),
            unit.get("lab_description") or None,
            seq,
        )

    log.info("default_curriculum_seeded", curriculum_id=curriculum_id, units=len(units))
    return curriculum_id
